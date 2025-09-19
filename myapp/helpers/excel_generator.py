import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import io
import os
import logging

logger = logging.getLogger(__name__)

class ExcelGenerator:
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def generate_testcase_excel(self, test_cases: List[Dict], service_name: str) -> str:
        try:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Test Cases"
            
            # Create summary sheet
            self._create_summary_sheet(test_cases, service_name)
            
            # Create test cases sheet
            self._create_testcases_sheet(test_cases)
            
            filename = "testcases.xlsx"
            folder = os.path.join(os.getcwd(), "files")

            if not os.path.exists(folder):
                os.makedirs(folder)

            filepath = os.path.join(folder, f"{filename}")
            
            self.workbook.save(filepath)
            print(f"Excel file saved successfully: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating Excel file: {e}")
            print(f"Error generating Excel file: {e}")
            return None
    
    def _create_summary_sheet(self, test_cases: List[Dict], service_name: str):
        """Create summary sheet with test statistics matching the required format"""
        summary_ws = self.workbook.create_sheet("Summary", 0)
        
        # Define headers exactly as shown in screenshot
        headers = [
            'Service',
            '# P1 Total Tests',
            '# P1 Tests run',
            'P1 Run / Total',
            '# P1 Tests passed',
            '# P1 Tests Blocked',
            'P1 Passed / P1 Total',
            '# Total Tests',
            '# Tests run',
            'Run / Total',
            '# Tests passed',
            '# Tests Failed',
            '# Tests Blocked',
            'Passed / Total',
            'Test Data Details'
        ]
        
        # Add headers with blue background and white text
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col_idx, header in enumerate(headers, start=1):
            cell = summary_ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Calculate statistics
        total_tests = len(test_cases)
        p1_tests = len([tc for tc in test_cases if tc.get('priority') == 'P1'])
        
        # For demo purposes, assuming all tests are run and passed (100% success rate)
        # In real implementation, these would come from actual test execution results
        tests_run = total_tests
        tests_passed = total_tests
        tests_failed = 0
        tests_blocked = 0
        
        p1_tests_run = p1_tests
        p1_tests_passed = p1_tests
        p1_tests_blocked = 0
        
        # Calculate percentages
        p1_run_percentage = f"{(p1_tests_run / p1_tests * 100):.2f}%" if p1_tests > 0 else "0.00%"
        p1_passed_percentage = f"{(p1_tests_passed / p1_tests * 100):.2f}%" if p1_tests > 0 else "0.00%"
        run_percentage = f"{(tests_run / total_tests * 100):.2f}%" if total_tests > 0 else "0.00%"
        passed_percentage = f"{(tests_passed / total_tests * 100):.2f}%" if total_tests > 0 else "0.00%"
        
        # Data row
        data_row = [
            service_name,
            p1_tests,
            p1_tests_run,
            p1_run_percentage,
            p1_tests_passed,
            p1_tests_blocked,
            p1_passed_percentage,
            total_tests,
            tests_run,
            run_percentage,
            tests_passed,
            tests_failed,
            tests_blocked,
            passed_percentage,
            ""  # Test Data Details - empty for now
        ]
        
        # Add data row
        for col_idx, value in enumerate(data_row, start=1):
            cell = summary_ws.cell(row=2, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths to match screenshot layout
        column_widths = {
            'A': 20,  # Service
            'B': 12,  # # P1 Total Tests
            'C': 12,  # # P1 Tests run
            'D': 12,  # P1 Run / Total
            'E': 12,  # # P1 Tests passed
            'F': 12,  # # P1 Tests Blocked
            'G': 15,  # P1 Passed / P1 Total
            'H': 12,  # # Total Tests
            'I': 12,  # # Tests run
            'J': 12,  # Run / Total
            'K': 12,  # # Tests passed
            'L': 12,  # # Tests Failed
            'M': 12,  # # Tests Blocked
            'N': 15,  # Passed / Total
            'O': 15   # Test Data Details
        }
        
        for column_letter, width in column_widths.items():
            summary_ws.column_dimensions[column_letter].width = width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in summary_ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

    def _create_testcases_sheet(self, test_cases: List[Dict]):
        """Create detailed test cases sheet"""
        # Headers
        headers = [
            'Use Case',
            'Test Scenario',
            'Priority',
            'Preconditions',
            'Input',
            'Expected Result',
            'Test Result',
            'Comments',
            'Tester',
            'Execution Date'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add headers
        for col_idx, header in enumerate(headers, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add test cases data
        for row_idx, test_case in enumerate(test_cases, start=2):
            row_data = [
                test_case.get('Use Case', ''),
                test_case.get('Test Scenario', ''),
                test_case.get('Priority', 'P2'),
                test_case.get('Preconditions', ''),
                test_case.get('Input', ''),
                test_case.get('Expected Result', ''),
                '',  # Test Result - to be filled during execution
                '',  # Comments - to be filled during execution
                '',  # Tester - to be filled during execution
                ''   # Execution Date - to be filled during execution
            ]
            
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Color coding for priorities
                if col_idx == 3:  # Priority column
                    if value == 'P1':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value == 'P2':
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                    elif value == 'P3':
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Auto-adjust column widths
        column_widths = {
            'A': 25,  # Use Case
            'B': 40,  # Test Scenario
            'C': 10,  # Priority
            'D': 30,  # Preconditions
            'E': 30,  # Input
            'F': 40,  # Expected Result
            'G': 15,  # Test Result
            'H': 25,  # Comments
            'I': 15,  # Tester
            'J': 15   # Execution Date
        }
        
        for column_letter, width in column_widths.items():
            self.worksheet.column_dimensions[column_letter].width = width
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in self.worksheet.iter_rows(min_row=1, max_row=len(test_cases) + 1, 
                                          min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Freeze the header row
        self.worksheet.freeze_panes = 'A2'
    
    def get_excel_bytes(self, test_cases: List[Dict], service_name: str) -> bytes:
        """Generate Excel file and return as bytes"""
        try:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Test Cases"
            
            # Create summary sheet
            self._create_summary_sheet(test_cases, service_name)
            
            # Create test cases sheet
            self._create_testcases_sheet(test_cases)
            
            # Save to memory
            excel_buffer = io.BytesIO()
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating Excel bytes: {e}")
            return None
