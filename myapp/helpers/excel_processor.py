import pandas as pd
from openpyxl import load_workbook
from typing import List, Dict, Any, Optional, Tuple
import logging
from django.core.files.uploadedfile import UploadedFile
from collections import defaultdict
import re

logger = logging.getLogger(__name__)

class ExcelProcessor:
    def __init__(self):
        self.test_case_columns = [
            'Use Case', 'Test Scenario', 'Priority', 'Preconditions', 
            'Input', 'Expected Result', 'Test Result', 'Comments', 
            'Tester', 'Execution Date'
        ]
        
        # Column mapping for different naming conventions
        self.column_mappings = {
            'use case': 'Use Case',
            'usecase': 'Use Case',
            'test scenario': 'Test Scenario',
            'scenario': 'Test Scenario',
            'test_scenario': 'Test Scenario',
            'priority': 'Priority',
            'preconditions': 'Preconditions',
            'precondition': 'Preconditions',
            'input': 'Input',
            'inputs': 'Input',
            'test data': 'Input',
            'expected result': 'Expected Result',
            'expected': 'Expected Result',
            'expected_result': 'Expected Result',
            'test result': 'Test Result',
            'result': 'Test Result',
            'test_result': 'Test Result',
            'status': 'Test Result',
            'comments': 'Comments',
            'comment': 'Comments',
            'remarks': 'Comments',
            'tester': 'Tester',
            'tested by': 'Tester',
            'execution date': 'Execution Date',
            'date': 'Execution Date',
            'executed on': 'Execution Date'
        }
    
    def extract_comprehensive_data_from_excel(self, excel_file: UploadedFile) -> Dict[str, Any]:
        """Extract comprehensive test data from Excel file"""
        try:
            # Load workbook
            workbook = load_workbook(excel_file, data_only=True)
            
            comprehensive_data = {
                'service_info': {},
                'test_cases': [],
                'summary_metrics': {},
                'raw_sheets_data': {}
            }
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Processing sheet: {sheet_name}")
                
                # Store raw sheet data
                comprehensive_data['raw_sheets_data'][sheet_name] = self._extract_sheet_data(sheet)
                
                # Identify sheet type and extract accordingly
                sheet_type = self._identify_sheet_type(sheet, sheet_name)
                
                if sheet_type == 'test_cases':
                    test_cases = self._extract_test_cases_from_sheet(sheet)
                    comprehensive_data['test_cases'].extend(test_cases)
            
            # Clean and validate data
            comprehensive_data = self._clean_and_validate_data(comprehensive_data)
            
            return comprehensive_data
            
        except Exception as e:
            logger.error(f"Error extracting comprehensive data from Excel: {e}")
            return {}
    
    def _extract_sheet_data(self, sheet) -> List[List[Any]]:
        """Extract raw data from a sheet"""
        try:
            data = []
            for row in sheet.iter_rows(values_only=True):
                # Convert None values to empty strings and clean data
                clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
                data.append(clean_row)
            return data
        except Exception as e:
            logger.error(f"Error extracting sheet data: {e}")
            return []
    
    def _identify_sheet_type(self, sheet, sheet_name: str) -> str:
        """Identify the type of sheet based on content and name"""
        try:
            sheet_type = None
            sheet_name_lower = sheet_name.lower()

            # Check content patterns
            sheet_text = self._get_sheet_text_content(sheet)
            
            # Check sheet name patterns
            if any(keyword in sheet_name_lower for keyword in ['statistic', 'summary', 'dashboard']):
                sheet_type = 'main_sheet'
            elif self._has_test_case_headers(sheet_text):
                sheet_type = 'test_cases'
            else:
                sheet_type = 'unknown'
            
            return sheet_type
            
        except Exception as e:
            logger.error(f"Error identifying sheet type: {e}")
            return 'unknown'
    
    def _get_sheet_text_content(self, sheet) -> str:
        try:
            text_content = ""
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text_content += str(cell) + " "
            return text_content.lower()
        except Exception as e:
            logger.error(f"Error getting sheet text content: {e}")
            return ""
    
    def _has_test_case_headers(self, text: str) -> bool:
        required_headers = ['use case', 'test scenario', 'priority']
        return sum(1 for header in required_headers if header in text) >= 2
    
    def _identify_sections_in_sheet(self, rows: List[Tuple]) -> Dict[str, Dict[str, int]]:
        try:
            sections = {}
            
            for i, row in enumerate(rows):
                row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
                
                # Look for title/header section
                if any(keyword in row_text for keyword in ['test execution report', 'test report', 'service']):
                    if 'title' not in sections:
                        sections['title'] = {'start': i, 'end': i + 3}
                
                # Look for statistics section
                elif any(keyword in row_text for keyword in ['test execution statistics', 'statistics', 'p1 total tests']):
                    sections['statistics'] = {'start': i, 'end': self._find_section_end(rows, i, 'statistics')}
                
                # Look for environment section
                elif any(keyword in row_text for keyword in ['test environment', 'environment', 'test data details']):
                    sections['environment'] = {'start': i, 'end': self._find_section_end(rows, i, 'environment')}
                
                # Look for test cases section
                elif any(keyword in row_text for keyword in ['detailed test cases', 'test cases', 'use case']):
                    if self._has_test_case_headers(' '.join([str(cell) for cell in row if cell])):
                        sections['test_cases'] = {'start': i, 'end': len(rows)}
            
            return sections
            
        except Exception as e:
            logger.error(f"Error identifying sections: {e}")
            return {}
    
    def _find_section_end(self, rows: List[Tuple], start: int, section_type: str) -> int:
        """Find the end of a section"""
        try:
            # Look for next major section or empty rows
            for i in range(start + 1, len(rows)):
                row_text = ' '.join([str(cell).lower() if cell else '' for cell in rows[i]])
                
                # Check for next major section
                if any(keyword in row_text for keyword in ['detailed test cases', 'test environment', 'statistics']):
                    return i
                
                # Check for multiple empty rows (section break)
                if i < len(rows) - 2:
                    next_rows_empty = all(
                        not any(cell for cell in rows[j] if cell)
                        for j in range(i, min(i + 3, len(rows)))
                    )
                    if next_rows_empty:
                        return i
            
            return min(start + 20, len(rows))  # Default section length
            
        except Exception as e:
            logger.error(f"Error finding section end: {e}")
            return start + 10
    
    def _extract_service_info_from_rows(self, rows: List[Tuple]) -> Dict[str, str]:
        """Extract service information from rows"""
        try:
            service_info = {}
            
            for row in rows:
                row_text = ' '.join([str(cell) if cell else '' for cell in row])
                
                # Look for service name patterns
                if 'test execution report' in row_text.lower():
                    # Extract service name after "Test Execution Report -"
                    match = re.search(r'test execution report\s*-\s*(.+)', row_text, re.IGNORECASE)
                    if match:
                        service_info['service_name'] = match.group(1).strip()
                
                # Look for service codes in brackets
                bracket_match = re.search(r'\[([^\]]+)\]', row_text)
                if bracket_match:
                    service_info['service_code'] = bracket_match.group(1)
            
            return service_info
            
        except Exception as e:
            logger.error(f"Error extracting service info from rows: {e}")
            return {}
    
    def _extract_test_cases_from_sheet(self, sheet) -> List[Dict[str, str]]:
        """Extract test cases from a dedicated test cases sheet"""
        try:
            # Convert to DataFrame for easier processing
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append([str(cell) if cell is not None else "" for cell in row])
            
            if not data:
                return []
            
            df = pd.DataFrame(data)
            return self._extract_test_cases_from_dataframe(df)
            
        except Exception as e:
            logger.error(f"Error extracting test cases from sheet: {e}")
            return []
    
    def _extract_test_cases_from_rows(self, rows: List[Tuple]) -> List[Dict[str, str]]:
        """Extract test cases from rows"""
        try:
            if not rows:
                return []
            
            # Convert to DataFrame
            data = [[str(cell) if cell is not None else "" for cell in row] for row in rows]
            df = pd.DataFrame(data)
            
            return self._extract_test_cases_from_dataframe(df)
            
        except Exception as e:
            logger.error(f"Error extracting test cases from rows: {e}")
            return []
    
    def _extract_test_cases_from_dataframe(self, df: pd.DataFrame) -> List[Dict[str, str]]:
        """Extract test cases from DataFrame"""
        try:
            test_cases = []
            
            # Find header row
            header_row_idx = self._find_header_row(df)
            if header_row_idx == -1:
                logger.warning("Could not find header row in test cases")
                return []
            
            # Extract headers
            headers = df.iloc[header_row_idx].tolist()
            headers = [str(h).strip() for h in headers if str(h).strip()]
            
            # Map headers to standard names
            mapped_headers = []
            for header in headers:
                clean_header = header.lower().strip()
                mapped_header = self.column_mappings.get(clean_header, header)
                mapped_headers.append(mapped_header)
            
            # Extract test case rows
            for idx in range(header_row_idx + 1, len(df)):
                row = df.iloc[idx].tolist()
                
                # Skip empty rows
                if not any(str(cell).strip() for cell in row if cell):
                    continue
                
                # Create test case dictionary
                test_case = {}
                for i, (header, value) in enumerate(zip(mapped_headers, row)):
                    if header and i < len(row):
                        clean_value = str(value).strip() if value else ""
                        test_case[header] = clean_value
                        
                # Only add if it has essential fields
                if (test_case.get('Use Case') or test_case.get('Test Scenario')) and test_case.get('Priority'):
                    test_cases.append(test_case)

            return test_cases
            
        except Exception as e:
            logger.error(f"Error extracting test cases from DataFrame: {e}")
            return []
    
    def _find_header_row(self, df: pd.DataFrame) -> int:
        """Find the row containing test case headers"""
        try:
            for idx, row in df.iterrows():
                row_text = ' '.join([str(cell).lower() for cell in row if cell])
                
                # Check for test case headers
                header_indicators = ['use case', 'test scenario', 'priority', 'expected result']
                matches = sum(1 for indicator in header_indicators if indicator in row_text)
                
                if matches >= 2:  # At least 2 header indicators
                    return idx
            
            return -1
            
        except Exception as e:
            logger.error(f"Error finding header row: {e}")
            return -1
    
    def _clean_and_validate_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Clean and validate extracted data"""
        try:
            # Clean test cases
            if 'test_cases' in data:
                cleaned_cases = []
                for tc in data['test_cases']:
                    cleaned_tc = {}
                    for key, value in tc.items():
                        # Clean value
                        if isinstance(value, str):
                            cleaned_value = re.sub(r'\s+', ' ', value.strip())
                            cleaned_value = cleaned_value.replace('\n', ' ').replace('\r', '')
                        else:
                            cleaned_value = str(value) if value else ""
                        
                        cleaned_tc[key] = cleaned_value
                    
                    # Only keep test cases with essential information
                    if cleaned_tc.get('Use Case') or cleaned_tc.get('Test Scenario'):
                        cleaned_cases.append(cleaned_tc)
                
                data['test_cases'] = cleaned_cases
            
            # Ensure all required fields exist
            if 'service_info' not in data:
                data['service_info'] = {}
            if 'test_statistics' not in data:
                data['test_statistics'] = {}
            if 'environment_info' not in data:
                data['environment_info'] = {}
            
            return data
            
        except Exception as e:
            logger.error(f"Error cleaning and validating data: {e}")
            return data
    
    def extract_text_representation(self, excel_file: UploadedFile) -> str:
        """Extract text representation of Excel file for embedding"""
        try:
            comprehensive_data = self.extract_comprehensive_data_from_excel(excel_file)
            
            text_parts = []
            
            # Add service information
            service_info = comprehensive_data.get('service_info', {})
            if service_info:
                text_parts.append("SERVICE INFORMATION:")
                for key, value in service_info.items():
                    text_parts.append(f"{key}: {value}")
                text_parts.append("")
            
            # Add test statistics
            stats = comprehensive_data.get('test_statistics', {})
            if stats:
                text_parts.append("TEST STATISTICS:")
                for key, value in stats.items():
                    text_parts.append(f"{key}: {value}")
                text_parts.append("")
            
            # Add test cases
            test_cases = comprehensive_data.get('test_cases', [])
            if test_cases:
                text_parts.append("TEST CASES:")
                for i, tc in enumerate(test_cases, 1):
                    text_parts.append(f"Test Case {i}:")
                    for key, value in tc.items():
                        if value:
                            text_parts.append(f"  {key}: {value}")
                    text_parts.append("")
            
            return "\n".join(text_parts)
            
        except Exception as e:
            logger.error(f"Error extracting text representation: {e}")
            return ""